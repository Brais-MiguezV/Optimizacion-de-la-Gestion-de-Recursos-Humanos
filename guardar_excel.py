import pandas as pd
from pathlib import Path
from sqlalchemy import create_engine, Engine, text
import os
import warnings
from dotenv import load_dotenv
from colorama import init, Fore, Style
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, datetime


# -------------------- Inicialización --------------------
warnings.filterwarnings("ignore", category=UserWarning)
load_dotenv()
init(autoreset=True)

# -------------------- Configuración de conexión --------------------
db_config = {
    "user": os.getenv("USUARIO", "test"),
    "password": os.getenv("PASS_USUARIO", "test"),
    "host": os.getenv("HOST", "postgres"),
    "port": os.getenv("PORT", 5432),
    "dbname": os.getenv("DATABASE", "Jira"),
}


def obtener_conexion() -> Engine:
    """
    Crea una conexión a la base de datos PostgreSQL utilizando SQLAlchemy.
    Returns:
        Engine: Un objeto Engine de SQLAlchemy para interactuar con la base de datos.
    """

    url = f"postgresql://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['dbname']}"  # Cadena de conexión para SQLAlchemy
    try:
        engine = create_engine(url)
        with engine.connect() as conn:
            pass  # Prueba de conexión
    except Exception as e:
        print(
            Fore.RED
            + Style.BRIGHT
            + f"\n❌ Error al conectar con la base de datos: {e}\n"
            + Style.RESET_ALL
        )
        exit(1)

    return engine


def cargar_datos(
    engine: Engine,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Carga los datos de tareas, empleados, proyectos e histórico de tareas desde la base de datos.
    Args:
        engine (Engine): El objeto Engine de SQLAlchemy para la conexión a la base de datos.
    Returns:
        tuple: Un tuple que contiene cuatro DataFrames:
            - tareas_mes_actual: Tareas del mes actual.
            - empleados: Información de los empleados.
            - proyectos: Información de los proyectos.
            - historico_tareas: Histórico de tareas con habilidades extraídas y candidatos.
    """

    hoy = date.today()  # Obtener la fecha actual
    primer_dia_mes = hoy.replace(day=1)  # Calcular el primer día del mes actual
    try:
        tareas_mes_actual = pd.read_sql(
            f"""
            SELECT
                t.id,
                p.proyecto || '-' || split_part(t.clave, '-', 2) AS clave,
                t.fecha,
                t.timespent_real,
                t.timespent_estimado,
                t.bien_estimado,
                p.proyecto AS nombre_proyecto,
                t.assignee_in_candidatos as "Empleado entre candidatos",
                a.empleado AS nombre_assignee,
                t.status_text,
                t.issue_type,
                t.texto,
                e.empleado AS nombre_candidato,
                eh.nivel_actual AS nivel_candidato,
                h.habilidad,
                LEAST(h.experiencia::numeric, 10) AS experiencia,
                t.fecha_modificacion
            FROM
                tareas t
                LEFT JOIN proyectos p ON p.codificacion = t.project_key
                LEFT JOIN empleados a ON a.codificacion = t.assignee
                LEFT JOIN LATERAL unnest(COALESCE(t.candidatos, ARRAY[]::candidatos[])) WITH ORDINALITY AS c(codificacion, porcentaje_acierto, fecha_modificacion, ord1) ON TRUE
                LEFT JOIN empleados e ON e.codificacion = c.codificacion
                LEFT JOIN LATERAL unnest(COALESCE(t.habilidades_extraidas, ARRAY[]::habilidades_tarea[])) WITH ORDINALITY AS h(habilidad, experiencia, fecha_modificacion, ord2) ON TRUE
                LEFT JOIN LATERAL (
                    SELECT eh.nivel_actual
                    FROM unnest(COALESCE(e.habilidades, ARRAY[]::habilidades_empleado[])) AS eh(habilidad, nivel_actual, fecha_modificacion)
                    WHERE eh.habilidad = h.habilidad
                    LIMIT 1
                ) eh ON TRUE
            WHERE
                (ord1 = ord2 OR (ord1 IS NULL AND ord2 IS NULL))
                AND (
                    t.fecha >= '{primer_dia_mes}'
                    OR (
                        t.status_text = 'In Progress'
                        AND t.fecha >= (date_trunc('month', CURRENT_DATE) - INTERVAL '3 months')
                    )
                )
            ORDER BY
                t.fecha asc;
            """,
            engine,
        )
        empleados = pd.read_sql(
            "SELECT id, empleado, is_active, habilidad.habilidad as habilidad, habilidad.nivel_actual, empleados.fecha_modificacion FROM empleados, unnest(habilidades) as habilidad order by id asc, habilidad desc",
            engine,
        )
        proyectos = pd.read_sql(
            """SELECT
                p.id,
                p.proyecto,
                h.habilidad,
                h.nivel_necesario,
                p.fecha_modificacion
            FROM
                proyectos p
                    LEFT JOIN LATERAL unnest(
                        COALESCE(p.habilidades_necesarias, ARRAY[]::habilidades_proyecto[])
                                    ) AS h(habilidad, nivel_necesario, fecha_modificacion)
                            ON TRUE
            ORDER BY
                p.id asc, habilidad desc""",
            engine,
        )
        historico_tareas = pd.read_sql(
            """
            SELECT
                t.id,
                p.proyecto || '-' || split_part(t.clave, '-', 2) AS clave,
                t.fecha,
                t.timespent_real,
                t.timespent_estimado,
                t.bien_estimado,
                p.proyecto AS nombre_proyecto,
                t.assignee_in_candidatos as "Empleado entre candidatos",
                a.empleado AS nombre_assignee,
                t.status_text,
                t.issue_type,
                t.texto,
                e.empleado AS nombre_candidato,
                eh.nivel_actual AS nivel_candidato,
                h.habilidad,
                LEAST(h.experiencia::numeric, 10) AS experiencia,
                t.fecha_modificacion
            FROM
                tareas t
                LEFT JOIN proyectos p ON p.codificacion = t.project_key
                LEFT JOIN empleados a ON a.codificacion = t.assignee
                LEFT JOIN LATERAL unnest(COALESCE(t.candidatos, ARRAY[]::candidatos[])) WITH ORDINALITY AS c(codificacion, porcentaje_acierto, fecha_modificacion, ord1) ON TRUE
                LEFT JOIN empleados e ON e.codificacion = c.codificacion
                LEFT JOIN LATERAL unnest(COALESCE(t.habilidades_extraidas, ARRAY[]::habilidades_tarea[])) WITH ORDINALITY AS h(habilidad, experiencia, fecha_modificacion, ord2) ON TRUE
                LEFT JOIN LATERAL (
                    SELECT eh.nivel_actual
                    FROM unnest(COALESCE(e.habilidades, ARRAY[]::habilidades_empleado[])) AS eh(habilidad, nivel_actual, fecha_modificacion)
                    WHERE eh.habilidad = h.habilidad
                    LIMIT 1
                ) eh ON TRUE
            WHERE
                (ord1 = ord2 OR (ord1 IS NULL AND ord2 IS NULL))
                order by t.fecha_modificacion asc
        """,
            engine,
        )
    except Exception as e:
        print(
            Fore.RED
            + Style.BRIGHT
            + f"\n❌ Error al cargar datos desde la base de datos: {e}\n"
            + Style.RESET_ALL
        )
        exit(1)

    print(
        f"{Fore.GREEN}✅ Datos de tareas, empleados, proyectos e histórico cargados correctamente\n{Style.RESET_ALL}"
    )
    return tareas_mes_actual, empleados, proyectos, historico_tareas


def color_por_estado(estado: str) -> str:
    """Devuelve un color en formato hexadecimal basado en el estado de la tarea.
    Esta función asigna un color específico a cada estado de tarea para su uso en hojas de cálculo Excel.
    Los colores son:
        - Resolved: Verde claro
        - Closed: Gris claro
        - In Progress: Amarillo
        - To Do: Rojo claro
    - Default: Blanco

    Args:
        estado (str): Estado de la tarea (Resolved, Closed, In Progress, To Do).

    Returns:
        str: Color en formato hexadecimal (ARGB) correspondiente al estado.
    """
    colores = {
        "Resolved": "FFC6EFCE",  # Verde claro
        "In Progress": "FFFFEB9C",  # Amarillo
        "To Do": "FFFFC7CE",  # Rojo claro
        "Open": "FFCCE5FF",  # Azul claro
        "Waiting": "FFB2B2B2",  # Gris medio
        "Closed": "abdbe3",  # Gris claro
        "Reopen": "FFCCE5FF",  # Azul claro
    }
    return colores.get(estado, "FFFFFFFF")  # Blanco por defecto


def aplicar_estilo_excel(writer: pd.ExcelWriter, hojas: list[str]):
    """
    Aplica estilos a las hojas de un archivo Excel.
    Args:
        writer (pd.ExcelWriter): El objeto ExcelWriter de pandas.
        hojas (list[str]): Lista de nombres de hojas a las que se aplicarán los estilos.
    """

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for nombre_hoja in hojas:
        ws = writer.sheets[nombre_hoja]

        for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ancho = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        header_font = Font(bold=True)
        alignment = Alignment(horizontal="center")
        fill = PatternFill(
            start_color="FF5F249F", end_color="FF5F249F", fill_type="solid"
        )
        header_font.color = "FCFCFC"

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = alignment
            cell.fill = fill
            cell.border = border

        header_cols = [cell.value for cell in ws[1]]
        if "status_text" in header_cols:
            col_estado_idx = header_cols.index("status_text") + 1

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                estado = row[col_estado_idx - 1].value
                fill_color = color_por_estado(estado)
                for cell in row:
                    cell.fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
                    cell.border = border
        else:
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, max_col=ws.max_column
            ):
                for cell in row:
                    cell.border = border


def generar_ruta_versionada(base_path: Path) -> Path:
    """
    Si el archivo ya existe, genera una nueva ruta con sufijo _v{n}.
    Args:
        base_path (Path): Ruta base sin versión.
    Returns:
        Path: Ruta disponible (nueva o versionada).
    """
    if not base_path.exists():
        return base_path

    version = 1
    while True:
        versioned = base_path.with_name(
            f"{base_path.stem}_v{version}{base_path.suffix}"
        )
        if not versioned.exists():
            return versioned
        version += 1


def guardar_base_datos_en_excel(
    tareas_mes_actual, empleados, proyectos, historico_tareas, ruta_salida
):
    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    print(f"{Fore.YELLOW}📝 Guardando base de datos en Excel{Style.RESET_ALL}")
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        hojas = {
            "Tareas Mes Actual o En Progreso": tareas_mes_actual,
            "Empleados": empleados,
            "Proyectos": proyectos,
            "Historico de tareas": historico_tareas,
        }

        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        aplicar_estilo_excel(writer, hojas.keys())

    print(
        f"{Fore.GREEN}\n✅  Base de datos guardada correctamente en: {ruta_salida}\n{Style.RESET_ALL}"
    )


if __name__ == "__main__":
    engine = obtener_conexion()
    tareas_mes_actual, empleados, proyectos, historico_tareas = cargar_datos(engine)
    nombre_base = f"gestion_{date.today().strftime('%Y%m%d')}.xlsx"
    ruta_base = Path("./salidas") / nombre_base
    ruta_final = generar_ruta_versionada(ruta_base)

    guardar_base_datos_en_excel(
        tareas_mes_actual,
        empleados,
        proyectos,
        historico_tareas,
        ruta_final,
    )
